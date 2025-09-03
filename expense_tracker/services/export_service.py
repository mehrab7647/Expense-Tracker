"""Export service for data export functionality."""

import csv
import os
from typing import List, Optional, Dict, Any
from datetime import date
from pathlib import Path
from decimal import Decimal

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from ..models.transaction import Transaction
from ..models.enums import TransactionType
from ..services.transaction_service import TransactionService
from ..services.report_service import ReportService


class ExportService:
    """Service for exporting financial data to various formats."""
    
    def __init__(self, transaction_service: TransactionService, report_service: Optional[ReportService] = None):
        """Initialize the export service.
        
        Args:
            transaction_service: Transaction service instance
            report_service: Report service instance (optional, for enhanced exports)
        """
        self.transaction_service = transaction_service
        self.report_service = report_service
    
    def export_transactions_to_csv(
        self,
        file_path: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        transaction_type: Optional[TransactionType] = None,
        category: Optional[str] = None
    ) -> bool:
        """Export transactions to CSV format.
        
        Args:
            file_path: Path where the CSV file will be saved
            start_date: Start date for filtering (optional)
            end_date: End date for filtering (optional)
            transaction_type: Filter by transaction type (optional)
            category: Filter by category (optional)
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            # Validate file path
            if not self._validate_file_path(file_path, '.csv'):
                return False
            
            # Get filtered transactions
            transactions = self._get_filtered_transactions(
                start_date, end_date, transaction_type, category
            )
            
            # Create directory if it doesn't exist
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            # Write CSV file
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = [
                    'ID', 'Date', 'Description', 'Category', 'Type', 'Amount'
                ]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                # Write header
                writer.writeheader()
                
                # Write transaction data
                for transaction in transactions:
                    writer.writerow({
                        'ID': transaction.id,
                        'Date': transaction.date.strftime('%Y-%m-%d') if transaction.date else '',
                        'Description': transaction.description,
                        'Category': transaction.category,
                        'Type': transaction.transaction_type.value,
                        'Amount': str(transaction.amount)
                    })
            
            return True
            
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return False
    
    def export_transactions_to_excel(
        self,
        file_path: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        transaction_type: Optional[TransactionType] = None,
        category: Optional[str] = None,
        include_summary: bool = True
    ) -> bool:
        """Export transactions to Excel format.
        
        Args:
            file_path: Path where the Excel file will be saved
            start_date: Start date for filtering (optional)
            end_date: End date for filtering (optional)
            transaction_type: Filter by transaction type (optional)
            category: Filter by category (optional)
            include_summary: Whether to include summary sheet
            
        Returns:
            True if export successful, False otherwise
        """
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl library not available. Install with: pip install openpyxl")
            return False
        
        try:
            # Validate file path
            if not self._validate_file_path(file_path, '.xlsx'):
                return False
            
            # Get filtered transactions
            transactions = self._get_filtered_transactions(
                start_date, end_date, transaction_type, category
            )
            
            # Create directory if it doesn't exist
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            # Create workbook
            workbook = openpyxl.Workbook()
            
            # Create transactions sheet
            self._create_transactions_sheet(workbook, transactions)
            
            # Create summary sheet if requested and report service is available
            if include_summary and self.report_service:
                self._create_summary_sheet(workbook, start_date, end_date, transaction_type)
            
            # Remove default sheet if we created others
            if len(workbook.worksheets) > 1 and 'Sheet' in [ws.title for ws in workbook.worksheets]:
                workbook.remove(workbook['Sheet'])
            
            # Save workbook
            workbook.save(file_path)
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    def export_category_summary_to_csv(
        self,
        file_path: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        transaction_type: Optional[TransactionType] = None
    ) -> bool:
        """Export category summary to CSV format.
        
        Args:
            file_path: Path where the CSV file will be saved
            start_date: Start date for filtering (optional)
            end_date: End date for filtering (optional)
            transaction_type: Filter by transaction type (optional)
            
        Returns:
            True if export successful, False otherwise
        """
        if not self.report_service:
            print("Error: Report service not available for category summary export")
            return False
        
        try:
            # Validate file path
            if not self._validate_file_path(file_path, '.csv'):
                return False
            
            # Get category breakdown report
            report = self.report_service.generate_category_breakdown_report(
                start_date, end_date, transaction_type
            )
            
            # Create directory if it doesn't exist
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            # Write CSV file
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = [
                    'Category', 'Total Amount', 'Transaction Count', 'Percentage', 'Average Amount'
                ]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                # Write header
                writer.writeheader()
                
                # Write category data
                for category, data in report['categories'].items():
                    writer.writerow({
                        'Category': category,
                        'Total Amount': str(data['total_amount']),
                        'Transaction Count': data['transaction_count'],
                        'Percentage': f"{data['percentage']:.2f}%",
                        'Average Amount': str(data['average_amount'])
                    })
            
            return True
            
        except Exception as e:
            print(f"Error exporting category summary to CSV: {e}")
            return False
    
    def export_monthly_report_to_excel(
        self,
        file_path: str,
        year: int
    ) -> bool:
        """Export monthly report to Excel format.
        
        Args:
            file_path: Path where the Excel file will be saved
            year: Year for the monthly report
            
        Returns:
            True if export successful, False otherwise
        """
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl library not available. Install with: pip install openpyxl")
            return False
        
        if not self.report_service:
            print("Error: Report service not available for monthly report export")
            return False
        
        try:
            # Validate file path
            if not self._validate_file_path(file_path, '.xlsx'):
                return False
            
            # Get monthly report
            report = self.report_service.generate_monthly_report(year)
            
            # Create directory if it doesn't exist
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            # Create workbook
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = f"Monthly Report {year}"
            
            # Create header
            headers = ['Month', 'Income', 'Expenses', 'Net Balance', 'Transaction Count']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add monthly data
            row = 2
            for month_name, data in report['monthly_data'].items():
                worksheet.cell(row=row, column=1, value=month_name)
                worksheet.cell(row=row, column=2, value=float(data['income']))
                worksheet.cell(row=row, column=3, value=float(data['expenses']))
                worksheet.cell(row=row, column=4, value=float(data['net_balance']))
                worksheet.cell(row=row, column=5, value=data['transaction_count'])
                row += 1
            
            # Add summary row
            row += 1
            worksheet.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
            worksheet.cell(row=row, column=2, value=float(report['summary']['total_income'])).font = Font(bold=True)
            worksheet.cell(row=row, column=3, value=float(report['summary']['total_expenses'])).font = Font(bold=True)
            worksheet.cell(row=row, column=4, value=float(report['summary']['net_balance'])).font = Font(bold=True)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            workbook.save(file_path)
            return True
            
        except Exception as e:
            print(f"Error exporting monthly report to Excel: {e}")
            return False
    
    def _get_filtered_transactions(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        transaction_type: Optional[TransactionType] = None,
        category: Optional[str] = None
    ) -> List[Transaction]:
        """Get filtered transactions based on criteria."""
        return self.transaction_service.filter_transactions(
            start_date=start_date,
            end_date=end_date,
            transaction_type=transaction_type,
            category=category
        )
    
    def _validate_file_path(self, file_path: str, expected_extension: str) -> bool:
        """Validate file path and extension.
        
        Args:
            file_path: Path to validate
            expected_extension: Expected file extension (e.g., '.csv', '.xlsx')
            
        Returns:
            True if valid, False otherwise
        """
        try:
            path = Path(file_path)
            
            # Check if path is absolute or relative
            if not path.is_absolute():
                # Convert to absolute path
                path = Path.cwd() / path
            
            # Check if parent directory exists or can be created
            parent_dir = path.parent
            if not parent_dir.exists():
                try:
                    parent_dir.mkdir(parents=True, exist_ok=True)
                except OSError:
                    print(f"Error: Cannot create directory {parent_dir}")
                    return False
            
            # Check file extension
            if not file_path.lower().endswith(expected_extension):
                print(f"Error: File must have {expected_extension} extension")
                return False
            
            # Check if we can write to the location
            try:
                # Try to create a temporary file to test write permissions
                test_file = path.with_suffix('.tmp')
                test_file.touch()
                test_file.unlink()
            except OSError:
                print(f"Error: No write permission for {path}")
                return False
            
            return True
            
        except Exception as e:
            print(f"Error validating file path: {e}")
            return False
    
    def _create_transactions_sheet(self, workbook: 'openpyxl.Workbook', transactions: List[Transaction]) -> None:
        """Create transactions sheet in Excel workbook."""
        worksheet = workbook.active
        worksheet.title = "Transactions"
        
        # Create header
        headers = ['ID', 'Date', 'Description', 'Category', 'Type', 'Amount']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add transaction data
        for row, transaction in enumerate(transactions, 2):
            worksheet.cell(row=row, column=1, value=transaction.id)
            worksheet.cell(row=row, column=2, value=transaction.date.strftime('%Y-%m-%d') if transaction.date else '')
            worksheet.cell(row=row, column=3, value=transaction.description)
            worksheet.cell(row=row, column=4, value=transaction.category)
            worksheet.cell(row=row, column=5, value=transaction.transaction_type.value)
            worksheet.cell(row=row, column=6, value=float(transaction.amount))
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(
        self,
        workbook: 'openpyxl.Workbook',
        start_date: Optional[date],
        end_date: Optional[date],
        transaction_type: Optional[TransactionType]
    ) -> None:
        """Create summary sheet in Excel workbook."""
        worksheet = workbook.create_sheet("Summary")
        
        # Get summary report
        summary = self.report_service.generate_summary_report(start_date, end_date)
        
        # Add title
        worksheet.cell(row=1, column=1, value="Financial Summary").font = Font(size=16, bold=True)
        
        # Add period information
        row = 3
        if start_date and end_date:
            worksheet.cell(row=row, column=1, value="Period:")
            worksheet.cell(row=row, column=2, value=f"{start_date} to {end_date}")
            row += 1
        
        if transaction_type:
            worksheet.cell(row=row, column=1, value="Type Filter:")
            worksheet.cell(row=row, column=2, value=transaction_type.value)
            row += 1
        
        # Add summary data
        row += 1
        summary_data = [
            ("Total Income", summary['totals']['total_income']),
            ("Total Expenses", summary['totals']['total_expenses']),
            ("Net Balance", summary['totals']['net_balance']),
            ("Total Transactions", summary['totals']['total_transactions']),
            ("", ""),
            ("Income Transactions", summary['counts']['income_transactions']),
            ("Expense Transactions", summary['counts']['expense_transactions']),
            ("", ""),
            ("Average Income", summary['averages']['average_income']),
            ("Average Expense", summary['averages']['average_expense']),
        ]
        
        for label, value in summary_data:
            if label:
                worksheet.cell(row=row, column=1, value=label).font = Font(bold=True)
                if isinstance(value, Decimal):
                    worksheet.cell(row=row, column=2, value=float(value))
                else:
                    worksheet.cell(row=row, column=2, value=value)
            row += 1
        
        # Auto-adjust column widths
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 15